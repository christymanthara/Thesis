import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


def save_results_to_excel(df_query, df_reference, accumulated_results, base_filename, 
                         current_comparison=None):
    """
    Save KNN results to a comprehensive Excel file with multiple sheets and formatting.
    
    Parameters:
    -----------
    df_query : pd.DataFrame
        DataFrame containing query transfer accuracy results
    df_reference : pd.DataFrame
        DataFrame containing reference CV accuracy results
    accumulated_results : dict
        Dictionary containing all accumulated results with metadata
    base_filename : str
        Base filename for the output Excel file
    current_comparison : str, optional
        The current source comparison being processed (for highlighting)
    
    Returns:
    --------
    str : Path to the saved Excel file
    """
    
    # Create output filename
    excel_filename = f"knn_results_comprehensive_{base_filename}.xlsx"
    
    try:
        # Create Excel writer object
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            
            # Sheet 1: Query Transfer Results
            df_query.to_excel(writer, sheet_name='Query Transfer', index=False)
            
            # Sheet 2: Reference CV Results  
            df_reference.to_excel(writer, sheet_name='Reference CV', index=False)
            
            # Sheet 3: Side-by-side comparison
            comparison_df = create_side_by_side_comparison(df_query, df_reference)
            comparison_df.to_excel(writer, sheet_name='Side by Side', index=False)
            
            # Sheet 4: Metadata summary
            metadata_df = create_metadata_dataframe(accumulated_results)
            metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
            
            # Sheet 5: Current run only (if specified)
            if current_comparison and current_comparison in accumulated_results:
                current_df = create_current_run_dataframe(accumulated_results, current_comparison)
                current_df.to_excel(writer, sheet_name='Current Run', index=False)
            
            # Sheet 6: Summary statistics
            summary_df = create_summary_statistics(accumulated_results)
            summary_df.to_excel(writer, sheet_name='Summary Stats', index=False)
        
        # Apply formatting to the Excel file
        format_excel_file(excel_filename, current_comparison, df_query, df_reference)
        
        print(f"Successfully saved comprehensive Excel results as {excel_filename}")
        return excel_filename
        
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        return None


def create_side_by_side_comparison(df_query, df_reference):
    """Create a side-by-side comparison of query and reference results."""
    if df_query.empty or df_reference.empty:
        return pd.DataFrame()
    
    # Get embedding columns (exclude 'Source Comparison')
    embedding_cols = [col for col in df_query.columns if col != 'Source Comparison']
    
    # Create comparison data
    comparison_data = {'Source Comparison': df_query['Source Comparison'].tolist()}
    
    for embedding in embedding_cols:
        comparison_data[f'{embedding}_Query'] = df_query[embedding].tolist()
        comparison_data[f'{embedding}_Reference'] = df_reference[embedding].tolist()
    
    return pd.DataFrame(comparison_data)


def create_current_run_dataframe(accumulated_results, current_comparison):
    """Create a DataFrame for the current run only."""
    if current_comparison not in accumulated_results:
        return pd.DataFrame()
    
    current_data = accumulated_results[current_comparison]
    query_results = current_data.get('query_accuracies', {})
    reference_results = current_data.get('reference_cv_scores', {})
    metadata = current_data.get('metadata', {})
    
    # Create DataFrame with current run data
    current_df_data = {
        'Metric': ['Source Comparison', 'Metadata'] + list(query_results.keys()) + 
                 [f'{key}_Reference' for key in reference_results.keys()],
        'Value': [current_comparison, metadata.get('metadata_display', 'No metadata')] + 
                list(query_results.values()) + list(reference_results.values()),
        'Type': ['Info', 'Info'] + ['Query Transfer'] * len(query_results) + 
                ['Reference CV'] * len(reference_results)
    }
    
    return pd.DataFrame(current_df_data)


def create_summary_statistics(accumulated_results):
    """Create summary statistics across all runs."""
    if not accumulated_results:
        return pd.DataFrame()
    
    # Collect all numeric results
    all_query_results = {}
    all_reference_results = {}
    
    for comparison_data in accumulated_results.values():
        query_results = comparison_data.get('query_accuracies', {})
        reference_results = comparison_data.get('reference_cv_scores', {})
        
        for embedding, accuracy in query_results.items():
            if embedding not in all_query_results:
                all_query_results[embedding] = []
            # Try to convert to float, skip if not possible
            try:
                if isinstance(accuracy, str):
                    # Remove any ± symbols and take first number
                    clean_acc = accuracy.split('±')[0].strip()
                    all_query_results[embedding].append(float(clean_acc))
                else:
                    all_query_results[embedding].append(float(accuracy))
            except (ValueError, TypeError):
                continue
        
        for embedding, accuracy in reference_results.items():
            if embedding not in all_reference_results:
                all_reference_results[embedding] = []
            try:
                if isinstance(accuracy, str):
                    clean_acc = accuracy.split('±')[0].strip()
                    all_reference_results[embedding].append(float(clean_acc))
                else:
                    all_reference_results[embedding].append(float(accuracy))
            except (ValueError, TypeError):
                continue
    
    # Calculate statistics
    summary_data = {
        'Embedding': [],
        'Query_Mean': [],
        'Query_Std': [],
        'Query_Min': [],
        'Query_Max': [],
        'Query_Count': [],
        'Reference_Mean': [],
        'Reference_Std': [],
        'Reference_Min': [],
        'Reference_Max': [],
        'Reference_Count': []
    }
    
    all_embeddings = set(all_query_results.keys()) | set(all_reference_results.keys())
    
    for embedding in sorted(all_embeddings):
        summary_data['Embedding'].append(embedding)
        
        # Query statistics
        query_vals = all_query_results.get(embedding, [])
        if query_vals:
            summary_data['Query_Mean'].append(f"{pd.Series(query_vals).mean():.3f}")
            summary_data['Query_Std'].append(f"{pd.Series(query_vals).std():.3f}")
            summary_data['Query_Min'].append(f"{min(query_vals):.3f}")
            summary_data['Query_Max'].append(f"{max(query_vals):.3f}")
            summary_data['Query_Count'].append(len(query_vals))
        else:
            for key in ['Query_Mean', 'Query_Std', 'Query_Min', 'Query_Max']:
                summary_data[key].append('N/A')
            summary_data['Query_Count'].append(0)
        
        # Reference statistics
        ref_vals = all_reference_results.get(embedding, [])
        if ref_vals:
            summary_data['Reference_Mean'].append(f"{pd.Series(ref_vals).mean():.3f}")
            summary_data['Reference_Std'].append(f"{pd.Series(ref_vals).std():.3f}")
            summary_data['Reference_Min'].append(f"{min(ref_vals):.3f}")
            summary_data['Reference_Max'].append(f"{max(ref_vals):.3f}")
            summary_data['Reference_Count'].append(len(ref_vals))
        else:
            for key in ['Reference_Mean', 'Reference_Std', 'Reference_Min', 'Reference_Max']:
                summary_data[key].append('N/A')
            summary_data['Reference_Count'].append(0)
    
    return pd.DataFrame(summary_data)


def create_metadata_dataframe(accumulated_results):
    """Create a DataFrame with metadata information (reused from your original code)."""
    metadata_data = {
        'Source Comparison': [],
        'Main Source': [],
        'Reference Source': [],
        'Main Tissue': [],
        'Reference Tissue': [],
        'Main Organism': [],
        'Reference Organism': [],
        'Metadata Display': []
    }
    
    for comparison, data in accumulated_results.items():
        metadata = data.get('metadata', {})
        metadata_data['Source Comparison'].append(comparison)
        metadata_data['Main Source'].append(metadata.get('main_source', 'unknown'))
        metadata_data['Reference Source'].append(metadata.get('ref_source', 'unknown'))
        metadata_data['Main Tissue'].append(metadata.get('main_tissue', 'unknown'))
        metadata_data['Reference Tissue'].append(metadata.get('ref_tissue', 'unknown'))
        metadata_data['Main Organism'].append(metadata.get('main_organism', 'unknown'))
        metadata_data['Reference Organism'].append(metadata.get('ref_organism', 'unknown'))
        metadata_data['Metadata Display'].append(metadata.get('metadata_display', 'No metadata'))
    
    return pd.DataFrame(metadata_data)


def format_excel_file(excel_filename, current_comparison, df_query, df_reference):
    """Apply formatting to the Excel file."""
    try:
        from openpyxl import load_workbook
        
        # Load the workbook
        wb = load_workbook(excel_filename)
        
        # Define styles
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        highlight_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)  # Cap at 30 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Format headers (first row)
            if ws.max_row > 0:
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = border
            
            # Apply borders and alignment to all cells
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
                    if cell.row > 1:  # Not header
                        cell.alignment = center_alignment
            
            # Highlight current comparison if applicable
            if current_comparison and sheet_name in ['Query Transfer', 'Reference CV']:
                df_to_check = df_query if sheet_name == 'Query Transfer' else df_reference
                if not df_to_check.empty and 'Source Comparison' in df_to_check.columns:
                    try:
                        current_row_idx = df_to_check[df_to_check['Source Comparison'] == current_comparison].index[0] + 2  # +2 for header and 0-indexing
                        for cell in ws[current_row_idx]:
                            cell.fill = highlight_fill
                    except (IndexError, KeyError):
                        pass  # Current comparison not found, skip highlighting
        
        # Save the formatted workbook
        wb.save(excel_filename)
        
    except ImportError:
        print("Warning: openpyxl not available for advanced formatting. Basic Excel file saved.")
    except Exception as e:
        print(f"Warning: Could not apply Excel formatting: {e}")


# Integration function to add to your existing workflow
def integrate_excel_export(df_query, df_reference, accumulated_results, base_filename, 
                          current_comparison=None):
    """
    Simple integration function that can be called from your existing create_results_table function.
    
    Add this line to your create_results_table function:
    integrate_excel_export(df_query, df_reference, accumulated_results, base_filename, source_comparison_key)
    """
    return save_results_to_excel(df_query, df_reference, accumulated_results, 
                                base_filename, current_comparison)


# Example usage
if __name__ == "__main__":
    # Example data structure (matches your existing format)
    accumulated_results_example = {
        'xin_2016 vs baron_2016h': {
            'query_accuracies': {'scVI': '0.834', 'scANVI': '0.867', 'RAW data': '0.598'},
            'reference_cv_scores': {'scVI': '0.850±0.023', 'scANVI': '0.901±0.015', 'RAW data': '0.612±0.045'},
            'metadata': {
                'main_source': 'xin_2016',
                'ref_source': 'baron_2016h',
                'main_tissue': 'pancreas',
                'ref_tissue': 'pancreas',
                'main_organism': 'human',
                'ref_organism': 'human',
                'metadata_display': 'Main: human_pancreas | Ref: human_pancreas'
            }
        }
    }
    
    # Create example DataFrames
    df_query_example = pd.DataFrame({
        'Source Comparison': ['xin_2016 vs baron_2016h'],
        'scVI': ['0.834'],
        'scANVI': ['0.867'],
        'RAW data': ['0.598']
    })
    
    df_reference_example = pd.DataFrame({
        'Source Comparison': ['xin_2016 vs baron_2016h'],
        'scVI': ['0.850±0.023'],
        'scANVI': ['0.901±0.015'],
        'RAW data': ['0.612±0.045']
    })
    
    # Test the function
    excel_file = save_results_to_excel(df_query_example, df_reference_example, 
                                      accumulated_results_example, "test", 
                                      'xin_2016 vs baron_2016h')
    print(f"Example Excel file created: {excel_file}")